import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Licences"

DARK  = "1A1A17"
GOLD  = "C8A96E"
LIGHT = "F5F0E8"
WHITE = "FFFFFF"

hdr_font = Font(bold=True, color=GOLD, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK)
hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin     = Side(style="thin", color="C8A96E")
brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    ("N°",            8),
    ("Date",         12),
    ("Nom Client",   22),
    ("Email",        30),
    ("Machine UUID", 42),
    ("Cle Licence",  24),
    ("Version",      10),
    ("Prix",         12),
    ("Statut",       14),
    ("Notes",        30),
]

for col, (title, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=title)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = hdr_aln
    cell.border    = brd
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 30

# Première licence déjà générée
row = [
    1,
    datetime.now().strftime("%d/%m/%Y"),
    "Client 1",
    "",
    "4C4C4544-0053-5210-8044-C7C04F5A4D32",
    "B6AD7-18D9B-9210F-DA7FF",
    "1.0.0",
    "199 EUR",
    "Actif",
    "",
]
fill_light = PatternFill("solid", fgColor=LIGHT)
for col, val in enumerate(row, 1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.fill      = fill_light
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = brd
    if col == 9:
        cell.font = Font(bold=True, color="27AE60")

# Lignes vides pré-formatées (15 lignes)
for r in range(3, 18):
    fill = PatternFill("solid", fgColor=WHITE if r % 2 == 0 else LIGHT)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col, value="")
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Onglet Résumé ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resume")
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 20

resume = [
    ("Total licences generees",  '=COUNTA(Licences!A2:A100)'),
    ("Licences actives",         '=COUNTIF(Licences!I2:I100,"Actif")'),
    ("Licences expirees",        '=COUNTIF(Licences!I2:I100,"Expire")'),
    ("Derniere mise a jour",     datetime.now().strftime("%d/%m/%Y %H:%M")),
]

for r, (label, val) in enumerate(resume, 1):
    c1 = ws2.cell(row=r, column=1, value=label)
    c1.font = Font(bold=True, color=DARK)
    c1.fill = PatternFill("solid", fgColor=LIGHT)
    c1.border = brd
    c1.alignment = Alignment(horizontal="left")

    c2 = ws2.cell(row=r, column=2, value=val)
    c2.font   = Font(bold=True, color=GOLD) if r < 4 else Font()
    c2.fill   = PatternFill("solid", fgColor=DARK) if r < 4 else PatternFill("solid", fgColor=LIGHT)
    c2.border = brd
    c2.alignment = Alignment(horizontal="center")

path = r"C:\Users\Smarte technologui\Desktop\Tableau_Licences_Amah.xlsx"
wb.save(path)
print(f"Tableau cree : {path}")
os.startfile(path)
