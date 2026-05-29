import sys
from pathlib import Path
from datetime import datetime

if getattr(sys, 'frozen', False):
    DEFAULT_PATH = str(Path(sys.executable).parent / "Documents")
else:
    DEFAULT_PATH = str(Path.home() / "Documents")


def read_excel(path: str) -> dict:
    """Lit un fichier Excel (.xlsx / .xls) et retourne son contenu."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl manquant — lance : pip install openpyxl"}

    p = Path(path).expanduser()
    if not p.exists():
        return {"error": f"Fichier introuvable : {path}"}

    try:
        wb     = openpyxl.load_workbook(str(p), data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws   = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    rows.append([str(c) if c is not None else "" for c in row])
            sheets[name] = rows[:100]
        return {"success": True, "fichier": p.name, "feuilles": sheets}
    except Exception as e:
        return {"error": str(e)}


def create_excel(filename: str, headers: list, rows: list,
                 save_path: str = None) -> dict:
    """Crée un fichier Excel avec en-têtes et données."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {"error": "openpyxl manquant — lance : pip install openpyxl"}

    try:
        save_dir = Path(save_path or DEFAULT_PATH)
        save_dir.mkdir(parents=True, exist_ok=True)
        filepath = save_dir / f"{filename}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données"

        # En-têtes stylisées (fond sombre, texte blanc, gras)
        hdr_fill   = PatternFill("solid", fgColor="1A1A17")
        hdr_font   = Font(bold=True, color="C8A96E", size=11)
        hdr_align  = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color="7a5f38")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = thin_border

        # Données
        for r_idx, row_data in enumerate(rows, 2):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            # Lignes alternées légèrement colorées
            if r_idx % 2 == 0:
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="F5F0E8")

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # Figer la première ligne
        ws.freeze_panes = "A2"

        wb.save(str(filepath))
        return {
            "success":  True,
            "fichier":  filepath.name,
            "chemin":   str(filepath),
            "lignes":   len(rows),
            "colonnes": len(headers),
        }
    except Exception as e:
        return {"error": str(e)}


def append_to_excel(path: str, rows: list) -> dict:
    """Ajoute des lignes à la fin d'un fichier Excel existant."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl manquant — lance : pip install openpyxl"}

    p = Path(path).expanduser()
    if not p.exists():
        return {"error": f"Fichier introuvable : {path}"}

    try:
        wb = openpyxl.load_workbook(str(p))
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(str(p))
        return {"success": True, "lignes_ajoutees": len(rows), "fichier": p.name}
    except Exception as e:
        return {"error": str(e)}
